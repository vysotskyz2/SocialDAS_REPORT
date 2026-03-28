from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


_CLR_PRIMARY = "2F5496"
_CLR_ACCENT = "4472C4"
_CLR_LIGHT = "D6E4F0"
_CLR_ZEBRA = "F2F7FB"
_CLR_KV_KEY = "E8EFF7"
_CLR_BORDER = "BDD0E7"
_CLR_WHITE = "FFFFFF"
_CLR_HEATMAP_LO = "F2F7FB"
_CLR_HEATMAP_HI = "2F5496"

_HEADER_FONT = Font(bold=True, color=_CLR_WHITE, size=10, name="Calibri")
_HEADER_FILL = PatternFill(start_color=_CLR_PRIMARY, end_color=_CLR_PRIMARY, fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_SECTION_FONT = Font(bold=True, color=_CLR_PRIMARY, size=12, name="Calibri")
_SECTION_FILL = PatternFill(start_color=_CLR_LIGHT, end_color=_CLR_LIGHT, fill_type="solid")
_SECTION_ALIGN = Alignment(vertical="center")

_KV_KEY_FONT = Font(bold=True, size=10, name="Calibri")
_KV_KEY_FILL = PatternFill(start_color=_CLR_KV_KEY, end_color=_CLR_KV_KEY, fill_type="solid")

_CELL_ALIGN = Alignment(vertical="center")
_CELL_FONT = Font(size=10, name="Calibri")
_ZEBRA_FILL = PatternFill(start_color=_CLR_ZEBRA, end_color=_CLR_ZEBRA, fill_type="solid")

_BORDER = Border(
    left=Side(style="thin", color=_CLR_BORDER),
    right=Side(style="thin", color=_CLR_BORDER),
    top=Side(style="thin", color=_CLR_BORDER),
    bottom=Side(style="thin", color=_CLR_BORDER),
)

_PCT_FMT = '0.00"%"'
_NUM_FMT = '#,##0'
_DEC_FMT = '#,##0.00'

_PCT_KEYS = frozenset({
    "avg_engagement_rate", "engagement_rate", "growth_rate_pct",
    "change_pct", "percentile",
})
_DEC_KEYS = frozenset({
    "slope", "r_squared", "intercept", "mean", "median", "std",
    "skewness", "kurtosis", "z_score", "composite_score",
    "sma_7", "sma_30", "ema_7", "overall_correlation",
    "avg_engagement", "avg_views", "avg_views_per_video",
    "avg_posts_per_week",
})
_INT_KEYS = frozenset({
    "followers", "following", "follows", "media_count",
    "total_likes", "total_comments", "total_views",
    "video_count", "subscriber_count", "view_count",
    "like_count", "comments_count", "comment_count",
    "share_count", "shares", "saves", "saved",
    "reach", "profile_views", "views", "website_clicks",
    "engagement", "follower_count", "following_count",
    "likes_count", "total_shares", "subscribers",
    "count", "min", "max", "total", "total_content",
    "day_offset", "projected_value", "value",
})


def _number_format_for_key(key: str) -> str | None:
    if key in _PCT_KEYS:
        return _PCT_FMT
    if key in _DEC_KEYS:
        return _DEC_FMT
    if key in _INT_KEYS:
        return _NUM_FMT
    return None


def _heatmap_fill(value: float, lo: float, hi: float) -> PatternFill:
    if hi == lo:
        ratio = 0.5
    else:
        ratio = max(0.0, min(1.0, (value - lo) / (hi - lo)))

    def _lerp(a: int, b: int) -> int:
        return int(a + (b - a) * ratio)

    r = _lerp(0xF2, 0x2F)
    g = _lerp(0xF7, 0x54)
    b = _lerp(0xFB, 0x96)
    clr = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill(start_color=clr, end_color=clr, fill_type="solid")


_OVERVIEW_FIELDS: dict[str, list[tuple[str, str]]] = {
    "instagram": [
        ("ID аккаунта", "account_id"),
        ("Имя пользователя", "username"),
        ("Подписчики", "followers"),
        ("Подписки", "following"),
        ("Публикации", "media_count"),
        ("Всего лайков", "total_likes"),
        ("Всего комментариев", "total_comments"),
        ("Средний ER (%)", "avg_engagement_rate"),
    ],
    "tiktok": [
        ("ID аккаунта", "account_id"),
        ("Имя", "display_name"),
        ("Подписчики", "followers"),
        ("Подписки", "following"),
        ("Всего лайков", "total_likes"),
        ("Кол-во видео", "video_count"),
        ("Среднее кол-во просмотров", "avg_views"),
        ("Средний ER (%)", "avg_engagement_rate"),
    ],
    "youtube": [
        ("ID аккаунта", "account_id"),
        ("Название канала", "title"),
        ("Подписчики", "subscribers"),
        ("Всего просмотров", "total_views"),
        ("Кол-во видео", "video_count"),
        ("Среднее просм./видео", "avg_views_per_video"),
        ("Средний ER (%)", "avg_engagement_rate"),
    ],
}

_FOLLOWERS_COLUMNS: dict[str, list[tuple[str, str]]] = {
    "instagram": [
        ("Дата", "date"),
        ("Подписчики", "followers"),
        ("Подписки", "follows"),
        ("Публикации", "media_count"),
    ],
    "tiktok": [
        ("Дата", "date"),
        ("Подписчики", "follower_count"),
        ("Подписки", "following_count"),
        ("Лайки", "likes_count"),
        ("Видео", "video_count"),
    ],
    "youtube": [
        ("Дата", "date"),
        ("Подписчики", "subscriber_count"),
        ("Видео", "video_count"),
        ("Просмотры", "view_count"),
    ],
}

_POSTS_COLUMNS: dict[str, list[tuple[str, str]]] = {
    "instagram": [
        ("ID", "ig_id"),
        ("Тип", "media_type"),
        ("Подпись", "caption"),
        ("Ссылка", "permalink"),
        ("Дата", "timestamp"),
        ("Лайки", "like_count"),
        ("Комментарии", "comments_count"),
        ("Вовлечённость", "engagement"),
        ("Охват", "reach"),
        ("Сохранения", "saved"),
        ("Просмотры", "views"),
        ("Репосты", "shares"),
    ],
    "tiktok": [
        ("ID", "tt_video_id"),
        ("Название", "title"),
        ("Длительность (сек)", "duration"),
        ("Ссылка", "share_url"),
        ("Дата", "create_time"),
        ("Лайки", "like_count"),
        ("Комментарии", "comment_count"),
        ("Репосты", "share_count"),
        ("Просмотры", "view_count"),
        ("ER (%)", "engagement_rate"),
    ],
    "youtube": [
        ("ID", "yt_video_id"),
        ("Название", "title"),
        ("Дата публикации", "published_at"),
        ("Длительность", "duration"),
        ("Просмотры", "view_count"),
        ("Лайки", "like_count"),
        ("Комментарии", "comment_count"),
        ("ER (%)", "engagement_rate"),
    ],
}

_ENGAGEMENT_COLUMNS: dict[str, list[tuple[str, str]]] = {
    "instagram": [
        ("Дата", "date"),
        ("Период", "period"),
        ("Охват", "reach"),
        ("Просмотры профиля", "profile_views"),
        ("Просмотры", "views"),
        ("Лайки", "likes"),
        ("Комментарии", "comments"),
        ("Репосты", "shares"),
        ("Сохранения", "saves"),
        ("Клики на сайт", "website_clicks"),
    ],
    "tiktok": [
        ("Дата", "date"),
        ("Лайки", "total_likes"),
        ("Комментарии", "total_comments"),
        ("Репосты", "total_shares"),
        ("Просмотры", "total_views"),
        ("ER (%)", "engagement_rate"),
    ],
    "youtube": [
        ("Дата", "date"),
        ("Просмотры", "total_views"),
        ("Лайки", "total_likes"),
        ("Комментарии", "total_comments"),
        ("ER (%)", "engagement_rate"),
    ],
}

_REFERENCE_SECTIONS: list[tuple[str, list[tuple[str, str]]]] = [
    ("Общие метрики", [
        (
            "ER — Engagement Rate (коэффициент вовлечённости)",
            "Процент аудитории, которая активно взаимодействует с контентом "
            "(лайки + комментарии + репосты) относительно числа подписчиков.\n"
            "Пример: 500 лайков + 50 комментариев на пост у аккаунта с 10 000 подписчиков → ER = 5,5%.\n"
            "Хороший ER в Instagram: 1–3%, в TikTok: 4–8%.",
        ),
        (
            "Охват (Reach)",
            "Сколько уникальных пользователей увидело вашу публикацию.\n"
            "Пример: если 3 000 человек увидели ваш пост — охват = 3 000.",
        ),
        (
            "Просмотры (Views / Impressions)",
            "Общее число показов публикации, включая повторные.\n"
            "Один пользователь может дать несколько просмотров, поэтому просмотры ≥ охвата.",
        ),
    ]),
    ("Рост и тренды", [
        (
            "SMA — скользящее среднее (Simple Moving Average)",
            "Среднее значение метрики за последние N дней. Сглаживает ежедневные колебания.\n"
            "SMA-7 — среднее за 7 дней, SMA-30 — за 30 дней.\n"
            "Пример: подписчики за 7 дней: 100, 105, 98, 110, 107, 103, 112 → SMA-7 = 105.",
        ),
        (
            "EMA — экспоненциальное скользящее среднее",
            "Как SMA, но недавние данные получают больший вес.\n"
            "Реагирует на изменения быстрее, чем SMA.\n"
            "Если EMA-7 > SMA-7, значит в последние дни рост ускоряется.",
        ),
        (
            "Рост (%) — Growth Rate",
            "Процент изменения метрики за день: (сегодня − вчера) / вчера × 100%.\n"
            "Пример: вчера 1 000 подписчиков, сегодня 1 050 → рост = +5%.",
        ),
        (
            "Направление тренда (Direction)",
            "Показывает, растёт или падает метрика в целом: «up» (рост), "
            "«down» (спад) или «stable» (без изменений).",
        ),
    ]),
    ("Регрессионный анализ", [
        (
            "Наклон (Slope)",
            "Скорость изменения: на сколько единиц растёт (или падает) метрика за один день.\n"
            "Slope = 12.5 означает +12,5 подписчиков в день.\n"
            "Отрицательное значение — метрика снижается.",
        ),
        (
            "R² — коэффициент детерминации",
            "Насколько хорошо прямая линия описывает ваши данные (от 0 до 1).\n"
            "R² = 0.95 → 95% изменений объясняются трендом (очень хорошо).\n"
            "R² = 0.30 → данные «шумные», тренд слабый.",
        ),
        (
            "Intercept (точка пересечения)",
            "Начальное значение тренда при day=0. Используется для прогнозов.\n"
            "Прогноз = intercept + slope × день.",
        ),
        (
            "Прогноз (Projections)",
            "Оценка будущих значений на основе текущего тренда.\n"
            "Пример: slope = +10, intercept = 1000, day_offset = 7 → "
            "прогноз через 7 дней ≈ 1 070.",
        ),
    ]),
    ("Статистика контента", [
        (
            "Среднее (Mean)",
            "Сумма всех значений, делённая на их количество.\n"
            "Пример: лайки 50, 70, 30 → среднее = 50.",
        ),
        (
            "Медиана (Median)",
            "Значение «посередине», если упорядочить все числа по возрастанию.\n"
            "Пример: лайки 10, 50, 200 → медиана = 50. "
            "Лучше среднего показывает «типичное» значение.",
        ),
        (
            "Стандартное отклонение (Std Dev)",
            "Насколько сильно значения разбросаны вокруг среднего.\n"
            "Маленькое отклонение → стабильные результаты.\n"
            "Пример: лайки 48, 50, 52 → σ ≈ 2 (стабильно); 10, 50, 90 → σ ≈ 40 (нестабильно).",
        ),
        (
            "Асимметрия (Skewness)",
            "Показывает, смещено ли распределение влево или вправо.\n"
            "0 — симметрично; > 0 — много постов с низкими значениями, "
            "но есть «хиты» с высокими;\n< 0 — наоборот.",
        ),
        (
            "Эксцесс (Kurtosis)",
            "Описывает «остроту» распределения.\n"
            "Высокий эксцесс → много экстремальных значений (вирусные посты или провалы).\n"
            "Близок к 0 → результаты распределены равномерно.",
        ),
        (
            "Квартили (Q1–Q4)",
            "Разделение контента на 4 группы по 25%.\n"
            "Q1 — нижние 25% (слабый контент), Q4 — верхние 25% (лучший контент).\n"
            "Число в ячейке — количество постов в группе.",
        ),
    ]),
    ("Оценки публикаций", [
        (
            "Z-Score — отклонение от нормы",
            "Показывает, насколько пост отличается от «среднего» в стандартных отклонениях.\n"
            "Z = 0 — ровно средний пост.\n"
            "Z = +2 — сильно выше среднего (вирусный контент).\n"
            "Z = −2 — сильно ниже среднего (провальный пост).\n"
            "|Z| > 2 автоматически считается аномалией.",
        ),
        (
            "Перцентиль (Percentile)",
            "Какой процент других постов набрал меньше вовлечённости, чем данный.\n"
            "Пример: перцентиль 90 → данный пост лучше 90% всех остальных.",
        ),
        (
            "Составной балл (Composite Score)",
            "Комплексная оценка поста, учитывающая несколько метрик одновременно.\n"
            "Чем выше — тем лучше пост по совокупности показателей.",
        ),
        (
            "Аномалия (is_anomaly)",
            "Пост отмечается как аномальный, если его Z-Score по модулю > 2.\n"
            "True (Да) — пост резко отличается от остальных (как положительно, так и отрицательно).",
        ),
    ]),
    ("Паттерны публикаций", [
        (
            "Лучшее время (Best Time)",
            "День недели и час, когда ваши посты получают наибольшую вовлечённость.\n"
            "Рекомендуется публиковать контент именно в это время.",
        ),
        (
            "Тепловая карта (Heatmap)",
            "Таблица 7 дней × 24 часа. Значение — средняя вовлечённость постов, "
            "опубликованных в данный день и час.\n"
            "Тёмные ячейки — высокая активность, светлые — низкая.",
        ),
        (
            "Среднее публикаций в неделю",
            "Общее число публикаций / количество недель в выбранном периоде.",
        ),
    ]),
    ("Корреляции", [
        (
            "Коэффициент корреляции (r)",
            "Связь между двумя метриками (от −1 до +1).\n"
            "r = +1 → идеально растут вместе.\n"
            "r = −1 → одна растёт, другая падает.\n"
            "r = 0 → связи нет.\n"
            "Пример: r(лайки, комментарии) = 0.85 → посты с большим числом лайков "
            "почти всегда собирают много комментариев.",
        ),
    ]),
    ("Сравнение периодов", [
        (
            "Изменение (%)",
            "На сколько процентов изменилась метрика между двумя периодами.\n"
            "Пример: среднее «до» = 200, «после» = 260 → изменение = +30%.",
        ),
        (
            "До / После (Before / After)",
            "Статистика за первую и вторую половину выбранного периода.\n"
            "Позволяет увидеть динамику: растёт ли аудитория, "
            "улучшается ли вовлечённость.",
        ),
    ]),
]


class ExcelBuilder:

    def build(
        self,
        platform: str,
        account_id: str,
        report_type: str,
        data: dict,
    ) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)

        builders = {
            "full": self._build_full,
            "overview": self._build_overview_only,
            "growth": self._build_growth_only,
            "engagement": self._build_engagement_only,
            "content": self._build_content_only,
        }
        builder = builders.get(report_type, self._build_full)
        builder(wb, platform, data)

        if not wb.sheetnames:
            ws = wb.create_sheet("Отчёт")
            ws.cell(row=1, column=1, value="Нет данных для формирования отчёта")

        self._sheet_reference(wb)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()


    def _build_full(self, wb: Workbook, platform: str, data: dict) -> None:
        if data.get("overview"):
            self._sheet_overview(wb, platform, data["overview"])
        if data.get("followers"):
            self._sheet_followers(wb, platform, data["followers"])
        if data.get("posts"):
            self._sheet_posts(wb, platform, data["posts"])
        if data.get("engagement"):
            self._sheet_engagement(wb, platform, data["engagement"])
        if data.get("growth"):
            self._sheet_growth(wb, data["growth"])
        if data.get("content_performance"):
            self._sheet_content_performance(wb, data["content_performance"])
        if data.get("posting_patterns"):
            self._sheet_posting_patterns(wb, data["posting_patterns"])
        if data.get("trends"):
            self._sheet_trends(wb, data["trends"])

    def _build_overview_only(self, wb: Workbook, platform: str, data: dict) -> None:
        if data.get("overview"):
            self._sheet_overview(wb, platform, data["overview"])

    def _build_growth_only(self, wb: Workbook, platform: str, data: dict) -> None:
        if data.get("growth"):
            self._sheet_growth(wb, data["growth"])

    def _build_engagement_only(self, wb: Workbook, platform: str, data: dict) -> None:
        if data.get("engagement"):
            self._sheet_engagement(wb, platform, data["engagement"])

    def _build_content_only(self, wb: Workbook, platform: str, data: dict) -> None:
        if data.get("content_performance"):
            self._sheet_content_performance(wb, data["content_performance"])

    def _sheet_overview(self, wb: Workbook, platform: str, overview: dict) -> None:
        ws = wb.create_sheet("Обзор")
        fields = _OVERVIEW_FIELDS.get(platform, _OVERVIEW_FIELDS["instagram"])
        self._write_section(ws, 1, "Обзор аккаунта", col_span=2)
        self._write_kv(ws, fields, overview, start_row=2)
        self._auto_width(ws)

    def _sheet_followers(self, wb: Workbook, platform: str, followers: dict) -> None:
        title = "Подписчики"
        ws = wb.create_sheet(title)
        columns = _FOLLOWERS_COLUMNS.get(platform, _FOLLOWERS_COLUMNS["instagram"])
        self._write_section(ws, 1, "Динамика подписчиков", col_span=len(columns))
        self._write_table(ws, columns, followers.get("data", []), start_row=2)
        ws.freeze_panes = "A3"
        self._auto_width(ws)

    def _sheet_posts(self, wb: Workbook, platform: str, posts: dict) -> None:
        title = "Публикации" if platform == "instagram" else "Видео"
        ws = wb.create_sheet(title)
        columns = _POSTS_COLUMNS.get(platform, _POSTS_COLUMNS["instagram"])
        rows = posts.get("data", [])
        total = posts.get("total_posts") or posts.get("total_videos")

        row = 1
        label = "публикаций" if platform == "instagram" else "видео"
        self._write_section(ws, row, f"Контент ({label})", col_span=len(columns))
        row += 1
        if total is not None:
            ws.cell(row=row, column=1, value=f"Всего: {total}")
            ws.cell(row=row, column=1).font = Font(bold=True, size=10, color=_CLR_PRIMARY, name="Calibri")
            row += 1

        self._write_table(ws, columns, rows, start_row=row)
        ws.freeze_panes = f"A{row + 1}"
        self._auto_width(ws)

    def _sheet_engagement(self, wb: Workbook, platform: str, engagement: dict) -> None:
        ws = wb.create_sheet("Вовлечённость")
        columns = _ENGAGEMENT_COLUMNS.get(platform, _ENGAGEMENT_COLUMNS["instagram"])
        self._write_section(ws, 1, "Метрики вовлечённости", col_span=len(columns))
        self._write_table(ws, columns, engagement.get("data", []), start_row=2)
        ws.freeze_panes = "A3"
        self._auto_width(ws)

    def _sheet_growth(self, wb: Workbook, growth: dict) -> None:
        ws = wb.create_sheet("Рост")
        row = 1

        reg = growth.get("regression", {})
        self._write_section(ws, row, "Регрессионный анализ", col_span=2)
        row += 1
        kv_fields = [
            ("Направление", "direction"),
            ("Наклон (Slope)", "slope"),
            ("R² (коэфф. детерминации)", "r_squared"),
            ("Intercept (начальное значение)", "intercept"),
        ]
        self._write_kv(ws, kv_fields, reg, start_row=row)
        row += len(kv_fields) + 1

        projections = growth.get("projections", [])
        if projections:
            row += 1
            self._write_section(ws, row, "Прогноз", col_span=3)
            row += 1
            proj_cols = [
                ("День", "day_offset"),
                ("Дата", "date"),
                ("Прогнозное значение", "projected_value"),
            ]
            self._write_table(ws, proj_cols, projections, start_row=row)
            row += len(projections) + 2

        data_points = growth.get("data", [])
        if data_points:
            row += 1
            self._write_section(ws, row, "Данные роста", col_span=6)
            row += 1
            growth_cols = [
                ("Дата", "date"),
                ("Значение", "value"),
                ("Рост (%)", "growth_rate_pct"),
                ("SMA-7 (среднее 7 дн.)", "sma_7"),
                ("SMA-30 (среднее 30 дн.)", "sma_30"),
                ("EMA-7 (экспон. среднее)", "ema_7"),
            ]
            self._write_table(ws, growth_cols, data_points, start_row=row)

        self._auto_width(ws)

    def _sheet_content_performance(self, wb: Workbook, content: dict) -> None:
        ws = wb.create_sheet("Контент")
        row = 1

        stats = content.get("engagement_stats", {})
        self._write_section(ws, row, "Статистика вовлечённости", col_span=2)
        row += 1
        stats_kv = [
            ("Количество постов", "count"),
            ("Среднее (Mean)", "mean"),
            ("Медиана", "median"),
            ("Стандартное откл. (σ)", "std"),
            ("Минимум", "min"),
            ("Максимум", "max"),
            ("Асимметрия (Skewness)", "skewness"),
            ("Эксцесс (Kurtosis)", "kurtosis"),
        ]
        self._write_kv(ws, stats_kv, stats, start_row=row)
        row += len(stats_kv) + 1

        quartiles = content.get("quartile_distribution", {})
        if quartiles:
            row += 1
            self._write_section(ws, row, "Распределение по квартилям", col_span=2)
            row += 1
            q_fields = [
                ("Q1 — нижние 25%", "q1"),
                ("Q2 — 25–50%", "q2"),
                ("Q3 — 50–75%", "q3"),
                ("Q4 — лучшие 25%", "q4"),
            ]
            self._write_kv(ws, q_fields, quartiles, start_row=row)
            row += len(q_fields) + 1

        items = content.get("items", [])
        if items:
            row += 1
            self._write_section(ws, row, "Оценки публикаций", col_span=6)
            row += 1
            item_cols = [
                ("ID контента", "content_id"),
                ("Вовлечённость", "engagement"),
                ("Перцентиль (%)", "percentile"),
                ("Z-Score (откл.)", "z_score"),
                ("Составной балл", "composite_score"),
                ("Аномалия", "is_anomaly"),
            ]
            self._write_table(ws, item_cols, items, start_row=row)

        self._auto_width(ws)

    def _sheet_posting_patterns(self, wb: Workbook, patterns: dict) -> None:
        ws = wb.create_sheet("Паттерны")
        row = 1

        self._write_section(ws, row, "Активность публикаций", col_span=3)
        row += 1
        summary_kv = [
            ("Всего публикаций", "total_content"),
            ("Среднее публикаций в неделю", "avg_posts_per_week"),
        ]
        self._write_kv(ws, summary_kv, patterns, start_row=row)
        row += len(summary_kv) + 1

        best = patterns.get("best_time")
        if best:
            row += 1
            self._write_section(ws, row, "Лучшее время для публикации", col_span=3)
            row += 1
            best_kv = [
                ("День недели", "day"),
                ("Час", "hour"),
                ("Средняя вовлечённость", "avg_engagement"),
            ]
            self._write_kv(ws, best_kv, best, start_row=row)
            row += len(best_kv) + 1

        by_day = patterns.get("by_day_of_week", [])
        if by_day:
            row += 1
            self._write_section(ws, row, "Вовлечённость по дням недели", col_span=3)
            row += 1
            day_cols = [
                ("День", "day"),
                ("№ дня", "day_index"),
                ("Средняя вовлечённость", "avg_engagement"),
            ]
            self._write_table(ws, day_cols, by_day, start_row=row)
            row += len(by_day) + 2

        by_hour = patterns.get("by_hour", [])
        if by_hour:
            row += 1
            self._write_section(ws, row, "Вовлечённость по часам", col_span=2)
            row += 1
            hour_cols = [("Час", "hour"), ("Средняя вовлечённость", "avg_engagement")]
            self._write_table(ws, hour_cols, by_hour, start_row=row)
            row += len(by_hour) + 2

        heatmap = patterns.get("heatmap", [])
        if heatmap:
            row += 1
            self._write_section(ws, row, "Тепловая карта (день × час)", col_span=25)
            row += 1

            all_vals: list[float] = []
            for day_data in heatmap:
                if isinstance(day_data, list):
                    for v in day_data:
                        if isinstance(v, (int, float)):
                            all_vals.append(float(v))
            lo = min(all_vals) if all_vals else 0
            hi = max(all_vals) if all_vals else 1

            ws.cell(row=row, column=1, value="День / Час")
            ws.cell(row=row, column=1).font = _HEADER_FONT
            ws.cell(row=row, column=1).fill = _HEADER_FILL
            ws.cell(row=row, column=1).alignment = _HEADER_ALIGN
            ws.cell(row=row, column=1).border = _BORDER
            for h in range(24):
                cell = ws.cell(row=row, column=h + 2, value=f"{h:02d}:00")
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _HEADER_ALIGN
                cell.border = _BORDER
            row += 1

            days_ru = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
            for i, day_data in enumerate(heatmap):
                day_label = days_ru[i] if i < len(days_ru) else f"День {i}"
                day_cell = ws.cell(row=row, column=1, value=day_label)
                day_cell.font = Font(bold=True, size=10, name="Calibri")
                day_cell.fill = _KV_KEY_FILL
                day_cell.border = _BORDER
                if isinstance(day_data, list):
                    for j, val in enumerate(day_data):
                        cell = ws.cell(row=row, column=j + 2, value=val)
                        cell.border = _BORDER
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.font = _CELL_FONT
                        if isinstance(val, (int, float)):
                            cell.fill = _heatmap_fill(float(val), lo, hi)
                            if float(val) > (lo + hi) / 2:
                                cell.font = Font(size=10, name="Calibri", color=_CLR_WHITE)
                row += 1

        self._auto_width(ws)

    def _sheet_trends(self, wb: Workbook, trends: dict) -> None:
        ws = wb.create_sheet("Тренды")
        row = 1

        reg = trends.get("regression", {})
        self._write_section(ws, row, "Анализ трендов", col_span=2)
        row += 1
        trend_kv = [
            ("Направление", "direction"),
            ("Наклон (Slope)", "slope"),
            ("R² (коэфф. детерминации)", "r_squared"),
        ]
        self._write_kv(ws, trend_kv, reg, start_row=row)
        row += len(trend_kv) + 1

        comparison = trends.get("period_comparison")
        if comparison:
            row += 1
            self._write_section(ws, row, "Сравнение периодов", col_span=2)
            row += 1
            change_kv = [("Изменение (%)", "change_pct")]
            self._write_kv(ws, change_kv, comparison, start_row=row)
            row += 2

            for period_name, label in [("before", "До"), ("after", "После")]:
                period = comparison.get(period_name, {})
                if period:
                    row += 1
                    self._write_section(ws, row, f"Период «{label}»", col_span=2)
                    row += 1
                    period_kv = [
                        ("Среднее", "mean"),
                        ("Медиана", "median"),
                        ("Всего", "total"),
                        ("Количество", "count"),
                    ]
                    self._write_kv(ws, period_kv, period, start_row=row)
                    row += len(period_kv) + 1

        anomalies = trends.get("anomalies", [])
        if anomalies:
            row += 1
            self._write_section(ws, row, "Аномалии", col_span=4)
            row += 1
            anom_cols = [
                ("ID контента", "content_id"),
                ("Дата", "date"),
                ("Значение", "value"),
                ("Z-Score (откл.)", "z_score"),
            ]
            self._write_table(ws, anom_cols, anomalies, start_row=row)
            row += len(anomalies) + 2

        correlations = trends.get("correlations", [])
        if correlations:
            row += 1
            self._write_section(ws, row, "Корреляции между метриками", col_span=3)
            row += 1
            corr_cols = [
                ("Метрика A", "metric_a"),
                ("Метрика B", "metric_b"),
                ("Коэффициент (r)", "overall_correlation"),
            ]
            self._write_table(ws, corr_cols, correlations, start_row=row)

        self._auto_width(ws)


    def _sheet_reference(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Справка")
        row = 1

        # title
        self._write_section(ws, row, "Справка — описание метрик отчёта", col_span=2)
        row += 2
        ws.cell(row=row, column=1, value=(
            "В этой вкладке описаны все метрики, используемые в отчёте. "
            "Если какой-то термин непонятен — найдите его здесь."
        ))
        ws.cell(row=row, column=1).font = Font(italic=True, size=10, name="Calibri", color="555555")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 2

        for section_title, items in _REFERENCE_SECTIONS:
            self._write_section(ws, row, section_title, col_span=2)
            row += 1

            for ci, hdr in enumerate(["Метрика", "Описание и пример"], 1):
                cell = ws.cell(row=row, column=ci, value=hdr)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _HEADER_ALIGN
                cell.border = _BORDER
            row += 1

            for idx, (metric, description) in enumerate(items):
                is_odd = idx % 2 == 0
                for ci, val in enumerate([metric, description], 1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.font = Font(
                        bold=(ci == 1), size=10, name="Calibri",
                    )
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = _BORDER
                    if is_odd:
                        cell.fill = _ZEBRA_FILL
                row += 1
            row += 1

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 80

    @staticmethod
    def _write_section(
        ws,
        row: int,
        title: str,
        col_span: int = 1,
    ) -> None:
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = _SECTION_FONT
        cell.fill = _SECTION_FILL
        cell.alignment = _SECTION_ALIGN
        cell.border = _BORDER
        if col_span > 1:
            ws.merge_cells(
                start_row=row, start_column=1,
                end_row=row, end_column=col_span,
            )
            for c in range(2, col_span + 1):
                sc = ws.cell(row=row, column=c)
                sc.fill = _SECTION_FILL
                sc.border = _BORDER

    @staticmethod
    def _write_kv(
        ws,
        fields: list[tuple[str, str]],
        data: dict,
        start_row: int = 1,
    ) -> None:
        for i, (label, key) in enumerate(fields):
            r = start_row + i
            is_odd = i % 2 == 0

            label_cell = ws.cell(row=r, column=1, value=label)
            label_cell.font = _KV_KEY_FONT
            label_cell.fill = _KV_KEY_FILL
            label_cell.border = _BORDER
            label_cell.alignment = _CELL_ALIGN

            raw = data.get(key)
            value_cell = ws.cell(row=r, column=2, value=raw)
            value_cell.border = _BORDER
            value_cell.alignment = _CELL_ALIGN
            value_cell.font = _CELL_FONT
            if is_odd:
                value_cell.fill = _ZEBRA_FILL

            fmt = _number_format_for_key(key)
            if fmt and isinstance(raw, (int, float)):
                value_cell.number_format = fmt

    @staticmethod
    def _write_table(
        ws,
        columns: list[tuple[str, str]],
        rows: list[dict],
        start_row: int = 1,
    ) -> None:
        for col_idx, (label, _) in enumerate(columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=label)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            cell.border = _BORDER

        for row_idx, row_data in enumerate(rows, start_row + 1):
            is_odd = (row_idx - start_row) % 2 == 1
            for col_idx, (_, key) in enumerate(columns, 1):
                raw = row_data.get(key) if isinstance(row_data, dict) else None
                cell = ws.cell(row=row_idx, column=col_idx, value=raw)
                cell.border = _BORDER
                cell.alignment = _CELL_ALIGN
                cell.font = _CELL_FONT
                if is_odd:
                    cell.fill = _ZEBRA_FILL

                fmt = _number_format_for_key(key)
                if fmt and isinstance(raw, (int, float)):
                    cell.number_format = fmt

    @staticmethod
    def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
        for col_cells in ws.columns:
            length = min_width
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    val_len = max(len(line) for line in str(cell.value).split("\n")) + 2
                    length = max(length, min(val_len, max_width))
            ws.column_dimensions[col_letter].width = length
