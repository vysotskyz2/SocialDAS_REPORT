from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


CLR_PRIMARY = "2F5496"
CLR_ACCENT = "4472C4"
CLR_LIGHT = "D6E4F0"
CLR_ZEBRA = "F2F7FB"
CLR_KV_KEY = "E8EFF7"
CLR_BORDER = "BDD0E7"
CLR_WHITE = "FFFFFF"
CLR_HEATMAP_LO = "F2F7FB"
CLR_HEATMAP_HI = "2F5496"

HEADER_FONT = Font(bold=True, color=CLR_WHITE, size=10, name="Calibri")
HEADER_FILL = PatternFill(start_color=CLR_PRIMARY, end_color=CLR_PRIMARY, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SECTION_FONT = Font(bold=True, color=CLR_PRIMARY, size=12, name="Calibri")
SECTION_FILL = PatternFill(start_color=CLR_LIGHT, end_color=CLR_LIGHT, fill_type="solid")
SECTION_ALIGN = Alignment(vertical="center")

KV_KEY_FONT = Font(bold=True, size=10, name="Calibri")
KV_KEY_FILL = PatternFill(start_color=CLR_KV_KEY, end_color=CLR_KV_KEY, fill_type="solid")

CELL_ALIGN = Alignment(vertical="center")
CELL_FONT = Font(size=10, name="Calibri")
ZEBRA_FILL = PatternFill(start_color=CLR_ZEBRA, end_color=CLR_ZEBRA, fill_type="solid")

BORDER = Border(
    left=Side(style="thin", color=CLR_BORDER),
    right=Side(style="thin", color=CLR_BORDER),
    top=Side(style="thin", color=CLR_BORDER),
    bottom=Side(style="thin", color=CLR_BORDER),
)

PCT_FMT = '0.00"%"'
NUM_FMT = '#,##0'
DEC_FMT = '#,##0.00'

PCT_KEYS = frozenset({
    "avg_engagement_rate", "engagement_rate", "growth_rate_pct",
    "change_pct", "percentile",
})
DEC_KEYS = frozenset({
    "slope", "r_squared", "intercept", "mean", "median", "std",
    "skewness", "kurtosis", "z_score", "composite_score",
    "sma_7", "sma_30", "ema_7", "overall_correlation",
    "avg_engagement", "avg_views", "avg_views_per_video",
    "avg_posts_per_week",
})
INT_KEYS = frozenset({
    "followers", "following", "follows", "media_count",
    "total_likes", "total_comments", "total_views",
    "video_count", "subscriber_count", "view_count",
    "like_count", "comments_count", "comment_count",
    "share_count", "shares", "saves", "saved",
    "reach", "profile_views", "views", "website_clicks",
    "engagement", "follower_count", "following_count",
    "likes_count", "total_shares", "subscribers",
    "count", "min", "max", "total", "total_content",
    "day_offset", "projected_value", "value",
})

OVERVIEW_FIELDS: dict[str, list[tuple[str, str]]] = {
    "instagram": [
        ("ID аккаунта", "account_id"),
        ("Имя пользователя", "username"),
        ("Подписчики", "followers"),
        ("Подписки", "following"),
        ("Публикации", "media_count"),
        ("Всего лайков", "total_likes"),
        ("Всего комментариев", "total_comments"),
        ("Средний ER (%)", "avg_engagement_rate"),
    ],
    "tiktok": [
        ("ID аккаунта", "account_id"),
        ("Имя", "display_name"),
        ("Подписчики", "followers"),
        ("Подписки", "following"),
        ("Всего лайков", "total_likes"),
        ("Кол-во видео", "video_count"),
        ("Среднее кол-во просмотров", "avg_views"),
        ("Средний ER (%)", "avg_engagement_rate"),
    ],
    "youtube": [
        ("ID аккаунта", "account_id"),
        ("Название канала", "title"),
        ("Подписчики", "subscribers"),
        ("Всего просмотров", "total_views"),
        ("Кол-во видео", "video_count"),
        ("Среднее просм./видео", "avg_views_per_video"),
        ("Средний ER (%)", "avg_engagement_rate"),
    ],
}

FOLLOWERS_COLUMNS: dict[str, list[tuple[str, str]]] = {
    "instagram": [
        ("Дата", "date"),
        ("Подписчики", "followers"),
        ("Подписки", "follows"),
        ("Публикации", "media_count"),
    ],
    "tiktok": [
        ("Дата", "date"),
        ("Подписчики", "follower_count"),
        ("Подписки", "following_count"),
        ("Лайки", "likes_count"),
        ("Видео", "video_count"),
    ],
    "youtube": [
        ("Дата", "date"),
        ("Подписчики", "subscriber_count"),
        ("Видео", "video_count"),
        ("Просмотры", "view_count"),
    ],
}

POSTS_COLUMNS: dict[str, list[tuple[str, str]]] = {
    "instagram": [
        ("ID", "ig_id"),
        ("Тип", "media_type"),
        ("Подпись", "caption"),
        ("Ссылка", "permalink"),
        ("Дата", "timestamp"),
        ("Лайки", "like_count"),
        ("Комментарии", "comments_count"),
        ("Вовлечённость", "engagement"),
        ("Охват", "reach"),
        ("Сохранения", "saved"),
        ("Просмотры", "views"),
        ("Репосты", "shares"),
    ],
    "tiktok": [
        ("ID", "tt_video_id"),
        ("Название", "title"),
        ("Длительность (сек)", "duration"),
        ("Ссылка", "share_url"),
        ("Дата", "create_time"),
        ("Лайки", "like_count"),
        ("Комментарии", "comment_count"),
        ("Репосты", "share_count"),
        ("Просмотры", "view_count"),
        ("ER (%)", "engagement_rate"),
    ],
    "youtube": [
        ("ID", "yt_video_id"),
        ("Название", "title"),
        ("Дата публикации", "published_at"),
        ("Длительность", "duration"),
        ("Просмотры", "view_count"),
        ("Лайки", "like_count"),
        ("Комментарии", "comment_count"),
        ("ER (%)", "engagement_rate"),
    ],
}

ENGAGEMENT_COLUMNS: dict[str, list[tuple[str, str]]] = {
    "instagram": [
        ("Дата", "date"),
        ("Период", "period"),
        ("Охват", "reach"),
        ("Просмотры профиля", "profile_views"),
        ("Просмотры", "views"),
        ("Лайки", "likes"),
        ("Комментарии", "comments"),
        ("Репосты", "shares"),
        ("Сохранения", "saves"),
        ("Клики на сайт", "website_clicks"),
    ],
    "tiktok": [
        ("Дата", "date"),
        ("Лайки", "total_likes"),
        ("Комментарии", "total_comments"),
        ("Репосты", "total_shares"),
        ("Просмотры", "total_views"),
        ("ER (%)", "engagement_rate"),
    ],
    "youtube": [
        ("Дата", "date"),
        ("Просмотры", "total_views"),
        ("Лайки", "total_likes"),
        ("Комментарии", "total_comments"),
        ("ER (%)", "engagement_rate"),
    ],
}

REFERENCE_SECTIONS: list[tuple[str, list[tuple[str, str]]]] = [
    ("Общие метрики", [
        (
            "ER — Engagement Rate (коэффициент вовлечённости)",
            "Процент аудитории, которая активно взаимодействует с контентом "
            "(лайки + комментарии + репосты) относительно числа подписчиков.\n"
            "Пример: 500 лайков + 50 комментариев на post у аккаунта с 10 000 подписчиков → ER = 5,5%.\n"
            "Хороший ER в Instagram: 1–3%, в TikTok: 4–8%.",
        ),
        (
            "Охват (Reach)",
            "Сколько уникальных пользователей увидело вашу публикацию.\n"
            "Пример: если 3 000 человек увидели ваш пост — охват = 3 000.",
        ),
        (
            "Просмотры (Views / Impressions)",
            "Общее число показов публикации, включая повторные.\n"
            "Один пользователь может дать несколько просмотров, поэтому просмотры ≥ охвата.",
        ),
    ]),
    ("Рост и тренды", [
        (
            "SMA — скользящее среднее (Simple Moving Average)",
            "Среднее значение метрики за последние N дней. Сглаживает ежедневные колебания.\n"
            "SMA-7 — среднее за 7 дней, SMA-30 — за 30 дней.\n"
            "Пример: подписчики за 7 дней: 100, 105, 98, 110, 107, 103, 112 → SMA-7 = 105.",
        ),
        (
            "EMA — экспоненциальное скользящее среднее",
            "Как SMA, но недавние данные получают больший вес.\n"
            "Реагирует на изменения быстрее, чем SMA.\n"
            "Если EMA-7 > SMA-7, значит в последние дни рост ускоряется.",
        ),
        (
            "Рост (%) — Growth Rate",
            "Процент изменения метрики за день: (сегодня − вчера) / вчера × 100%.\n"
            "Пример: вчера 1 000 подписчиков, сегодня 1 050 → рост = +5%.",
        ),
        (
            "Направление тренда (Direction)",
            "Показывает, растёт или падает метрика в целом: «up» (рост), "
            "«down» (спад) или «stable» (без изменений).",
        ),
    ]),
    ("Регрессионный анализ", [
        (
            "Наклон (Slope)",
            "Скорость изменения: на сколько единиц растёт (или падает) метрика за один день.\n"
            "Slope = 12.5 означает +12,5 подписчиков в день.\n"
            "Отрицательное значение — метрика снижается.",
        ),
        (
            "R² — коэффициент детерминации",
            "Насколько хорошо прямая линия описывает ваши данные (от 0 до 1).\n"
            "R² = 0.95 → 95% изменений объясняются трендом (очень хорошо).\n"
            "R² = 0.30 → данные «шумные», тренд слабый.",
        ),
        (
            "Intercept (точка пересечения)",
            "Начальное значение тренда при day=0. Используется для прогнозов.\n"
            "Прогноз = intercept + slope × день.",
        ),
        (
            "Прогноз (Projections)",
            "Оценка будущих значений на основе текущего тренда.\n"
            "Пример: slope = +10, intercept = 1000, day_offset = 7 → "
            "прогноз через 7 дней ≈ 1 070.",
        ),
    ]),
    ("Статистика контента", [
        (
            "Среднее (Mean)",
            "Сумма всех значений, делённая на их количество.\n"
            "Пример: лайки 50, 70, 30 → среднее = 50.",
        ),
        (
            "Медиана (Median)",
            "Значение «посередине», если упорядочить все числа по возрастанию.\n"
            "Пример: лайки 10, 50, 200 → медиана = 50. "
            "Лучше среднего показывает «типичное» значение.",
        ),
        (
            "Стандартное отклонение (Std Dev)",
            "Насколько сильно значения разбросаны вокруг среднего.\n"
            "Маленькое отклонение → стабильные результаты.\n"
            "Пример: лайки 48, 50, 52 → σ ≈ 2 (стабильно); 10, 50, 90 → σ ≈ 40 (нестабильно).",
        ),
        (
            "Асимметрия (Skewness)",
            "Показывает, смещено ли распределение влево или вправо.\n"
            "0 — симметрично; > 0 — много постов с низкими значениями, "
            "но есть «хиты» с высокими;\n< 0 — наоборот.",
        ),
        (
            "Эксцесс (Kurtosis)",
            "Описывает «остроту» распределения.\n"
            "Высокий эксцесс → много экстремальных значений (вирусные посты или провалы).\n"
            "Близок к 0 → результаты распределены равномерно.",
        ),
        (
            "Квартили (Q1–Q4)",
            "Разделение контента на 4 группы по 25%.\n"
            "Q1 — нижние 25% (слабый контент), Q2 — 25–50%, Q3 — 50–75%, Q4 — верхние 25% (лучший контент).\n"
            "Число в ячейке — количество постов в группе.",
        ),
    ]),
    ("Оценки публикаций", [
        (
            "Z-Score — отклонение от нормы",
            "Показывает, насколько пост отличается от «среднего» в стандартных отклонениях.\n"
            "Z = 0 — ровно средний пост.\n"
            "Z = +2 — сильно выше среднего (вирусный контент).\n"
            "Z = −2 — сильно ниже среднего (провальный пост).\n"
            "|Z| > 2 автоматически считается аномалией.",
        ),
        (
            "Перцентиль (Percentile)",
            "Какой процент других постов набрал меньше вовлечённости, чем данный.\n"
            "Пример: перцентиль 90 → данный пост лучше 90% всех остальных.",
        ),
        (
            "Составной балл (Composite Score)",
            "Комплексная оценка поста, учитывающая несколько метрик одновременно.\n"
            "Чем выше — тем лучше пост по совокупности показателей.",
        ),
        (
            "Аномалия (is_anomaly)",
            "Пост отмечается как аномальный, если его Z-Score по модулю > 2.\n"
            "True (Да) — пост резко отличается от остальных (как положительно, так и отрицательно).",
        ),
    ]),
    ("Паттерны публикаций", [
        (
            "Лучшее время (Best Time)",
            "День недели и час, когда ваши посты получают наибольшую вовлечённость.\n"
            "Рекомендуется публиковать контент именно в это время.",
        ),
        (
            "Тепловая карта (Heatmap)",
            "Таблица 7 дней × 24 часа. Значение — средняя вовлечённость постов, "
            "опубликованных в данный день и час.\n"
            "Тёмные ячейки — высокая активность, светлые — низкая.",
        ),
        (
            "Среднее публикаций в неделю",
            "Общее число публикаций / количество недель в выбранном периоде.",
        ),
    ]),
    ("Корреляции", [
        (
            "Коэффициент корреляции (r)",
            "Связь между двумя метриками (от −1 до +1).\n"
            "r = +1 → идеально растут вместе.\n"
            "r = −1 → одна растёт, другая падает.\n"
            "r = 0 → связи нет.\n"
            "Пример: r(лайки, комментарии) = 0.85 → посты с большим числом лайков "
            "почти всегда собирают много комментариев.",
        ),
    ]),
    ("Сравнение периодов", [
        (
            "Изменение (%)",
            "На сколько процентов изменилась метрика между двумя периодами.\n"
            "Пример: среднее «до» = 200, «после» = 260 → изменение = +30%.",
        ),
        (
            "До / После (Before / After)",
            "Статистика за первую и вторую половину выбранного периода.\n"
            "Позволяет увидеть динамику: растёт ли аудитория, "
            "улучшается ли вовлечённость.",
        ),
    ]),
]
